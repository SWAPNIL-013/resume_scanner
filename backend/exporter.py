import os
import pandas as pd
from typing import List, Dict
from datetime import datetime
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from utils import format_experience_years

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)


def get_new_excel_name(base_name="resumes", ext=".xlsx"):
    existing_files = sorted(glob.glob(os.path.join(EXPORTS_DIR, f"{base_name}*.xlsx")))
    if not existing_files:
        return os.path.join(EXPORTS_DIR, f"{base_name}.xlsx")
    else:
        nums = []
        for f in existing_files:
            fname = os.path.basename(f).replace(base_name, "").replace(ext, "")
            if fname.startswith("_") and fname[1:].isdigit():
                nums.append(int(fname[1:]))
        next_num = max(nums, default=1) + 1
        return os.path.join(EXPORTS_DIR, f"{base_name}_{next_num:02d}{ext}")


def multiline(text: str, sep=",") -> str:
    if not text:
        return ""
    parts = [p.strip() for p in text.replace(";", ",").split(",") if p.strip()]
    return "\n".join(parts)


def format_experience_list(experience: List[Dict]) -> str:
    lines = []
    for exp in experience:
        role = exp.get("role", "")
        company = exp.get("company", "")
        start = exp.get("start_date", "")
        end = exp.get("end_date", "")
        desc = exp.get("description", "")
        line = f"{role} at {company} ({start} - {end})"
        if desc:
            line += f"\n {desc}"
        lines.append(line)
    return "\n".join(lines)


def format_projects_list(projects: List[Dict]) -> str:
    lines = []
    for proj in projects:
        title = proj.get("title", "")
        techs = ", ".join(proj.get("technologies", []))
        desc = proj.get("description", "")
        line = f"{title} [{techs}]"
        if desc:
            line += f": {desc}"
        lines.append(line)
    return "\n".join(lines)


def write_resumes_to_sheet(ws, resume_list: List[Dict], is_new_sheet=False):
    start_row = ws.max_row + 1 if not is_new_sheet else 1

    if is_new_sheet:
        headers = [
            "Name", "Email", "Phone","Location", "URLs", "Skills",
            "Education", "Experience", "Projects", "Certifications",
            "Experience Years", "Score", "Remarks", "Uploaded At"
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        ws.freeze_panes = "A2"
        start_row = 2

    for resume in resume_list:
        row = [
            resume.get("name", ""),
            resume.get("email", ""),
            resume.get("phone", ""),
            resume.get("location",""),
            multiline("; ".join(resume.get("urls", []))) if isinstance(resume.get("urls"), list) else resume.get("urls", ""),
            multiline(", ".join(resume.get("skills", []))) if isinstance(resume.get("skills"), list) else resume.get("skills", ""),
            "; ".join(
                f"{edu.get('degree', '')} - {edu.get('institution', '')} ({edu.get('year', '')})"
                for edu in resume.get("education", []) if isinstance(edu, dict)
            ) if isinstance(resume.get("education"), list) else resume.get("education", ""),
            format_experience_list(resume.get("experience", [])) if isinstance(resume.get("experience"), list) else resume.get("experience", ""),
            format_projects_list(resume.get("projects", [])) if isinstance(resume.get("projects"), list) else resume.get("projects", ""),
            multiline("; ".join(resume.get("certifications", []))) if isinstance(resume.get("certifications"), list) else resume.get("certifications", ""),
            resume.get("total_experience_years", 0),
            resume.get("score", ""),
            multiline("; ".join(resume.get("remarks", []))) if isinstance(resume.get("remarks"), list) else resume.get("remarks", ""),
            resume.get("uploaded_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value:
                max_len = max(max_len, max(len(str(line)) for line in str(cell.value).split("\n")))
        ws.column_dimensions[col_letter].width = min(max_len + 5, 50)


def export_to_excel(
    resume_list: List[Dict],
    mode: str = "new_file",
    file_path: str = None,
    sheet_name: str = None
) -> pd.DataFrame:
    # Ensure file is in exports folder
    if file_path:
        file_path = os.path.join(EXPORTS_DIR, os.path.basename(file_path))
    else:
        file_path = get_new_excel_name()

    if mode == "new_file":
        wb = Workbook()
        ws = wb.create_sheet(title=sheet_name or "Sheet1")
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        write_resumes_to_sheet(ws, resume_list, is_new_sheet=True)

    elif mode == "append_sheet":
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in file.")
        ws = wb[sheet_name]
        write_resumes_to_sheet(ws, resume_list, is_new_sheet=False)

    elif mode == "new_sheet":
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists.")
        ws = wb.create_sheet(title=sheet_name)
        write_resumes_to_sheet(ws, resume_list, is_new_sheet=True)

    else:
        raise ValueError(f"Unknown mode: {mode}")

    wb.save(file_path)
    print(f"✅ Exported to {file_path} [{mode}] → sheet: {sheet_name or 'Sheet1'}")

    return pd.DataFrame(resume_list)
