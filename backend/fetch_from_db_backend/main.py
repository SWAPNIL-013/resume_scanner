import json
import shutil
import tempfile
from fastapi import APIRouter, FastAPI, UploadFile, File, Query, Body, Header, HTTPException, status
from typing import List

from openpyxl import load_workbook
from shared.parser import extract_text_from_jd
from shared.schema import ExportRequest, LoginRequest, RegisterRequest
from shared.auth import register_user, authenticate_user, create_access_token, get_user_from_token
from shared.exporter import export_to_excel, get_new_excel_name, EXPORTS_DIR,export_to_mongo
from shared.llm import generate_jd_json
import base64
import os
from shared.pipeline import run_pipeline_db
from pymongo import MongoClient



router=APIRouter()


@router.post("/register")
async def register(user: RegisterRequest):
    created = register_user(
        user.username,
        user.password,
        user.full_name,
        role="user"
    )

    if not created:
        raise HTTPException(status_code=400, detail="User already exists")

    return {"status": "success", "message": "User registered, waiting for admin approval"}


@router.post("/login")
async def login(user: LoginRequest):
    user_data = authenticate_user(user.username, user.password)

    if not user_data:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    if user_data.get("error") == "not_approved":
        raise HTTPException(status_code=403, detail="Admin approval pending")

    if user_data.get("error") == "denied":
        raise HTTPException(status_code=403, detail="Your request was denied")


    token = create_access_token({
        "username": user_data["username"],
        "role": user_data.get("role", "user")
    })

    return {"access_token": token, "token_type": "bearer"}




# --------------------------
# 1️⃣ Connect to MongoDB
# --------------------------
@router.post("/connect_mongo")
async def connect_mongo(
    mongo_url: str = Body(..., description="MongoDB connection URI"),
    db_name: str = Body(..., description="Database name"),
    collection_name: str = Body(..., description="Collection name"),
    authorization: str = Header(None)
):
    if not authorization:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing Authorization header")
    token = authorization.split(" ")[-1]
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired token")

    try:
        client = MongoClient(mongo_url)
        db = client[db_name]
        collection = db[collection_name]
        resume_count = collection.count_documents({})
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))
    print(f"Connection successful : Resume Count = {resume_count}")
    return {"status": "success", "resume_count": resume_count}



@router.post("/upload_jd")
async def upload_jd(
    file: UploadFile = File(...),
    authorization: str = Header(None),
    x_model: str = Header(None),
    x_api_key: str = Header(None)
):
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")

    token = authorization.split(" ")[-1]
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    model = x_model or "gemini-2.5-flash"
    api_key = x_api_key

    # ✅ Create temporary file for JD
    with tempfile.NamedTemporaryFile(delete=False, suffix=file.filename) as temp_file:
        shutil.copyfileobj(file.file, temp_file)
        temp_path = temp_file.name

    try:
        # ✅ Extract JD text
        jd_text = extract_text_from_jd(temp_path)

        # ✅ Convert JD → JD JSON
        jd_json = generate_jd_json(jd_text, api_key=api_key, model=model)
        print(json.dumps(jd_json, indent=2))

        # ✅ Extract ONLY JD fields for slider creation
        jd_fields = list(jd_json.keys())

        return {
            "status": "success",
            "jd_fields": jd_fields,
            "jd_json": jd_json
        }

    finally:
        # ✅ Always cleanup temp file
        try:
            os.remove(temp_path)
        except:
            pass




# --------------------------
# 3️⃣ Evaluate Resumes
# --------------------------
@router.post("/evaluate_resumes_db")
async def evaluate_resumes_db(
    mongo_url: str = Body(...),
    db_name: str = Body(...),
    collection_name: str = Body(...),
    jd_data: dict = Body(None),
    authorization: str = Header(None),
    x_model: str = Header(None),
    x_api_key: str = Header(None),
):
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    token = authorization.split(" ")[-1]
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    model = x_model or (jd_data and jd_data.get("model")) or "gemini-2.5-flash"
    api_key = x_api_key or (jd_data and jd_data.get("api_key"))

    jd_json = jd_data.get("jd_json") if jd_data else None
    weights = jd_data.get("weights") if jd_data else None

    if jd_json:
        print("✅ JD JSON received — scoring enabled\n")
    else:
        print("⚙️ No JD provided — try again")

    try:
        processed_resumes = run_pipeline_db(
            mongo_url=mongo_url,
            db_name=db_name,
            collection_name=collection_name,
            weights=weights,
            jd_json=jd_json,
            username=user.get("username"),
            api_key=api_key,
            model=model,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Pipeline processing failed: {e}")

    if not processed_resumes:
        return {"status": "error", "message": "No valid resumes to process."}

    return {
        "status": "success",
        "count": len(processed_resumes),
        "data": processed_resumes,
        "jd_mode": "enabled" if jd_data and jd_data.get("jd_text") else "disabled",
    }




# --------------------------
# 1️⃣ List user export files API
# --------------------------
@router.get("/list_exports")
async def list_exports(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing Authorization header")
    token = authorization.split(" ")[-1]
    user = get_user_from_token(token)
    if not user or not user.get("username"):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired token")

    user_base = os.path.join(EXPORTS_DIR, user["username"])
    if not os.path.exists(user_base):
        return {"files": []}

    files = [f for f in os.listdir(user_base) if f.endswith(".xlsx")]
    return {"files": files}

# --------------------------
# 2️⃣ List sheets in a user export file API
# --------------------------
@router.get("/list_sheets")
async def list_sheets(
    file_name: str = Query(..., description="Excel filename in user exports folder"),
    authorization: str = Header(None)
):
    if not authorization:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing Authorization header")
    token = authorization.split(" ")[-1]
    user = get_user_from_token(token)
    if not user or not user.get("username"):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired token")

    user_base = os.path.join(EXPORTS_DIR, user["username"])
    file_path = os.path.join(user_base, file_name)

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return {"sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read sheets: {str(e)}")

# --------------------------
# 3️⃣ Export resumes to Excel endpoint (your existing code)
# --------------------------
@router.post("/export_resumes_excel")
async def export_resumes_excel(req: ExportRequest, authorization: str = Header(None), x_model: str = Header(None), x_api_key: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing Authorization header")
    token = authorization.split(" ")[-1]
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired token")

    try:
        user_base = None
        if user and user.get("username"):
            user_base = os.path.join(EXPORTS_DIR, user.get("username"))
            os.makedirs(user_base, exist_ok=True)

        if req.file_path:
            # req.file_path here is just filename from frontend (e.g. "resumes.xlsx")
            req.file_path = os.path.join(user_base if user_base else EXPORTS_DIR, os.path.basename(req.file_path))
        else:
            # fallback if no file_path provided
            req.file_path = get_new_excel_name(base_dir=user_base)

        df = export_to_excel(
            resume_list=req.processed_resumes,
            mode=req.mode,
            file_path=req.file_path,
            sheet_name=req.sheet_name,
            base_dir=user_base
        )

        with open(req.file_path, "rb") as f:
            excel_bytes = f.read()
            excel_b64 = base64.b64encode(excel_bytes).decode("utf-8")

        return {
            "status": "success",
            "count": len(req.processed_resumes),
            "excel_file": excel_b64,
            "saved_path": req.file_path
        }

    except FileNotFoundError as e:
        return {"status": "error", "message": str(e)}
    except ValueError as e:
        return {"status": "error", "message": str(e)}
    except Exception as e:
        return {"status": "error", "message": f"Unexpected error: {str(e)}"}
# --------------------------
# 5️⃣ Export to MongoDB
# --------------------------
@router.post("/export_resumes_mongo")
async def export_resumes_mongo(req: dict, authorization: str = Header(None)):
    """
    req = {
        "processed_resumes": [...],
        "mongo_url": "mongodb+srv://user:pass@cluster/dbname",
        "db_name": "resume_db",
        "collection_name": "resumes"
    }
    """
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")

    try:
        mongo_url = req.get("mongo_url")
        db_name = req.get("db_name", "resume_db")
        collection_name = req.get("collection_name", "resumes")
        resume_list = req.get("processed_resumes", [])

        if not mongo_url:
            raise ValueError("MongoDB URL missing")

        result = export_to_mongo(resume_list, mongo_url, db_name, collection_name)
        print("✅ Data export received for MongoDB")
        return {
            "status": "success",
            "inserted_count": result["inserted_count"],
            "db_name": db_name,
            "collection_name": collection_name
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}



