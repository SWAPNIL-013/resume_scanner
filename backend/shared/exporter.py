import os
import pandas as pd
from typing import List, Dict
from datetime import datetime
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from pymongo import MongoClient
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
EXPORTS_DIR = PROJECT_ROOT / "exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

def get_new_excel_name(base_name="resumes", ext=".xlsx", base_dir: str = None):
    """
    Return a new excel filename inside base_dir (defaults to EXPORTS_DIR).
    """
    if base_dir is None:
        base_dir = EXPORTS_DIR
    os.makedirs(base_dir, exist_ok=True)
    existing_files = sorted(glob.glob(os.path.join(base_dir, f"{base_name}*.xlsx")))
    if not existing_files:
        return os.path.join(base_dir, f"{base_name}.xlsx")
    else:
        nums = []
        for f in existing_files:
            fname = os.path.basename(f).replace(base_name, "").replace(ext, "")
            if fname.startswith("_") and fname[1:].isdigit():
                nums.append(int(fname[1:]))
        next_num = max(nums, default=1) + 1
        return os.path.join(base_dir, f"{base_name}_{next_num:02d}{ext}")


def multiline(text: str, sep=",") -> str:
    if not text:
        return ""
    parts = [p.strip() for p in text.replace(";", ",").split(",") if p.strip()]
    return "\n".join(parts)

def format_experience_list(experience: List[Dict]) -> str:
    blocks = []

    for idx, exp in enumerate(experience, start=1):
        company = exp.get("company", "")
        start = exp.get("start_date", "")
        end = exp.get("end_date", "")
        years = exp.get("experience_years", exp.get("duration_years", ""))

        block = [
            f"{idx}) Company Name : {company}",
            f"   Duration     : {start} - {end}",
            f"   Exp Years    : {years}",
        ]

        blocks.append("\n".join(block))

    return "\n\n".join(blocks)

def format_projects_list(projects: List[Dict]) -> str:
    blocks = []

    for idx, proj in enumerate(projects, start=1):
        title = proj.get("title", "")
        techs = ", ".join(proj.get("technologies", []))
        desc = proj.get("description", "")

        block = [
            f"{idx}) Title : {title}",
            f"   Tech  : {techs}",
        ]

        if desc:
            block.append(f"   Desc  : {desc}")

        blocks.append("\n".join(block))

    return "\n\n".join(blocks)

def format_education_list(education: List[Dict]) -> str:
    blocks = []

    for idx, edu in enumerate(education, start=1):
        degree = edu.get("degree", "")
        inst = edu.get("institution", "")
        year = edu.get("year", "")

        block = [
            f"{idx}) Degree      : {degree}",
            f"   Institution : {inst}",
            f"   Year        : {year}",
        ]

        blocks.append("\n".join(block))

    return "\n\n".join(blocks)

def format_certifications_list(certifications: List[str]) -> str:
    return "\n".join(
        f"{idx}) {cert}"
        for idx, cert in enumerate(certifications, start=1)
        if cert
    )



# def write_resumes_to_sheet(ws, resume_list: List[Dict], is_new_sheet=False):
#     start_row = ws.max_row + 1 if not is_new_sheet else 1

#     if is_new_sheet:
  
#         headers = [
#             "Name", "Email", "Phone", "Location", "URLs", "Skills",
#             "Education", "Experience", "Projects", "Certifications",
#             "Total Experience Years",
#             "Matched Skills", "Missing Skills", "Other Skills",
#             "Score", "Remarks", "Uploaded At"
#         ]


#         ws.append(headers)
#         for col in range(1, len(headers) + 1):
#             ws.cell(row=1, column=col).font = Font(bold=True)
#         ws.freeze_panes = "A2"
#         start_row = 2

#     for resume in resume_list:
#         row = [
#             resume.get("name", ""),
#             resume.get("email", ""),
#             resume.get("phone", ""),
#             resume.get("location",""),
#             multiline("; ".join(resume.get("urls", []))) if isinstance(resume.get("urls"), list) else resume.get("urls", ""),
#             multiline(", ".join(resume.get("skills", []))) if isinstance(resume.get("skills"), list) else resume.get("skills", ""),
#             format_education_list(resume.get("education", []))if isinstance(resume.get("education"), list)else resume.get("education", ""),
#             format_experience_list(resume.get("experience", [])) if isinstance(resume.get("experience"), list) else resume.get("experience", ""),
#             format_projects_list(resume.get("projects", [])) if isinstance(resume.get("projects"), list) else resume.get("projects", ""),
#             format_certifications_list(resume.get("certifications", []))if isinstance(resume.get("certifications"), list)else resume.get("certifications", ""),
#             resume.get("total_experience_years", 0),
#             multiline(", ".join(resume.get("matched_skills", []))),
#             multiline(", ".join(resume.get("missing_skills", []))),
#             multiline(", ".join(resume.get("other_skills", []))),
#             resume.get("score", ""),
#             "\n".join(
#                 f"{idx}) {remark}"
#                 for idx, remark in enumerate(resume.get("remarks", []), start=1)
#             ) if isinstance(resume.get("remarks"), list) else resume.get("remarks", ""),

#             resume.get("uploaded_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),

#         ]
#         ws.append(row)

#     # Adjust column widths
#     for col in ws.columns:
#         max_len = 0
#         col_letter = col[0].column_letter
#         for cell in col:
#             cell.alignment = Alignment(wrap_text=True, vertical="top")
#             if cell.value:
#                 max_len = max(max_len, max(len(str(line)) for line in str(cell.value).split("\n")))
#         ws.column_dimensions[col_letter].width = min(max_len + 5, 50)

def write_resumes_to_sheet(ws, documents: List[Dict], is_new_sheet=False):
    start_row = ws.max_row + 1 if not is_new_sheet else 1

    if is_new_sheet:
        headers = [
            "Name", "Email", "Phone", "Location", "URLs", "Skills",
            "Education", "Experience", "Projects", "Certifications",
            "Total Experience Years",
            "Matched Skills", "Missing Skills", "Other Skills",
            "Score", "Remarks", "Uploaded At"
        ]

        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        ws.freeze_panes = "A2"
        start_row = 2

    for doc in documents:
        resume=doc.get("resume_json",{})
        # Get latest evaluation from evaluations list if exists
        evaluations = doc.get("evaluations", [])
        latest_eval = evaluations[-1] if evaluations else {}

        row = [
            resume.get("name", ""),
            resume.get("email", ""),
            resume.get("phone", ""),
            resume.get("location", ""),
            multiline("; ".join(resume.get("urls", []))) if isinstance(resume.get("urls"), list) else resume.get("urls", ""),
            multiline(", ".join(resume.get("skills", []))) if isinstance(resume.get("skills"), list) else resume.get("skills", ""),
            format_education_list(resume.get("education", [])) if isinstance(resume.get("education"), list) else resume.get("education", ""),
            format_experience_list(resume.get("experience", [])) if isinstance(resume.get("experience"), list) else resume.get("experience", ""),
            format_projects_list(resume.get("projects", [])) if isinstance(resume.get("projects"), list) else resume.get("projects", ""),
            format_certifications_list(resume.get("certifications", [])) if isinstance(resume.get("certifications"), list) else resume.get("certifications", ""),
            resume.get("total_experience_years", 0),

            multiline(", ".join(latest_eval.get("matched_skills", []))),
            multiline(", ".join(latest_eval.get("missing_skills", []))),
            multiline(", ".join(latest_eval.get("other_skills", []))),

            latest_eval.get("score", ""),

            "\n".join(
                f"{idx}) {overall_summary}"
                for idx, overall_summary in enumerate(latest_eval.get("overall_summary", []), start=1)
            ) if isinstance(latest_eval.get("overall_summary"), list) else latest_eval.get("overall_summary", ""),

            resume.get("uploaded_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value:
                max_len = max(max_len, max(len(str(line)) for line in str(cell.value).split("\n")))
        ws.column_dimensions[col_letter].width = min(max_len + 5, 50)


def export_to_excel(
    documents: List[Dict],
    mode: str = "new_file",
    file_path: str = None,
    sheet_name: str = None,
    base_dir: str = None
) -> pd.DataFrame:
    # Ensure file is in exports folder
    if base_dir is None:
        base_dir = EXPORTS_DIR

    if file_path:
        file_path = os.path.join(base_dir, os.path.basename(file_path))
    else:
        file_path = get_new_excel_name(base_dir=base_dir)

    if mode == "new_file":
        wb = Workbook()
        ws = wb.create_sheet(title=sheet_name or "Sheet1")
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        write_resumes_to_sheet(ws, documents, is_new_sheet=True)

    elif mode == "append_sheet":
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in file.")
        ws = wb[sheet_name]
        write_resumes_to_sheet(ws, documents, is_new_sheet=False)

    elif mode == "new_sheet":
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists.")
        ws = wb.create_sheet(title=sheet_name)
        write_resumes_to_sheet(ws, documents, is_new_sheet=True)

    else:
        raise ValueError(f"Unknown mode: {mode}")

    wb.save(file_path)
    print(f"✅ Exported to {file_path} [{mode}] → sheet: {sheet_name or 'Sheet1'}")

    return pd.DataFrame(documents)



# def export_to_mongo(
#     documents,
#     mongo_url,
#     db_name="resume_db",
#     collection_name="resumes"
# ):
#     if not mongo_url:
#         raise ValueError("MongoDB URL is required")

#     client = MongoClient(mongo_url)
#     db = client[db_name]
#     collection = db[collection_name]

#     # Unique identity
#     collection.create_index("resume_json.email", unique=True)

#     updated_count = 0

#     for doc in documents:
#         resume = doc.get("resume_json", {})
#         evaluations = doc.get("evaluations", [])
#         email = resume.get("email")
#         if not email:
#             continue

#         if not evaluations:
#             # Still ensure resume exists in DB
#             result = collection.update_one(
#                 {"resume_json.email": email},
#                 {
#                     "$setOnInsert": {
#                         "resume_json": resume,
#                         "evaluations": []
#                     }
#                 },
#                 upsert=True
#             )
#             if result.upserted_id:
#                 updated_count += 1
#             continue

#         # Append evaluations, resume_json only on insert
#         result = collection.update_one(
#             {"resume_json.email": email},
#             {
#                 "$setOnInsert": {
#                     "resume_json": resume
#                 },
#                 "$push": {
#                     "evaluations": {
#                         "$each": evaluations
#                     }
#                 }
#             },
#             upsert=True
#         )

#         if result.modified_count or result.upserted_id:
#             updated_count += 1

#     client.close()
#     return {"inserted_count": updated_count}

def export_to_mongo(
    documents,
    mongo_url,
    db_name="resume_db",
    collection_name="resumes"
):
    if not mongo_url:
        raise ValueError("MongoDB URL is required")

    client = MongoClient(mongo_url)
    db = client[db_name]
    collection = db[collection_name]

    # Unique identity
    collection.create_index("resume_json.email", unique=True)

    updated_count = 0

    for doc in documents:
        resume = doc.get("resume_json", {})
        evaluations = doc.get("evaluations", [])
        email = resume.get("email")
        if not email:
            continue

        if not evaluations:
            # Still ensure resume exists in DB
            result = collection.update_one(
                {"resume_json.email": email},
                {
                    "$set": {
                        "resume_json": resume
                    },
                    "$setOnInsert": {
                        "evaluations": []
                    }
                },
                upsert=True
            )
            if result.upserted_id or result.matched_count:
                updated_count += 1
            continue

        # Append evaluations, resume_json only on insert
        latest_eval=evaluations[-1]
        result = collection.update_one(
            {"resume_json.email": email},
            {
                "$setOnInsert": {
                    "resume_json": resume
                },
                "$push": {
                        "evaluations": latest_eval
                    }
                
            },
            upsert=True
        )

        if result.modified_count or result.upserted_id:
            updated_count += 1

    client.close()
    return {"inserted_count": updated_count}
