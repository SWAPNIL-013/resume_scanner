# import os
# import base64
# import requests
# import streamlit as st
# from openpyxl import load_workbook

# # --------------------------
# # Project Root
# # --------------------------
# PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# # --------------------------
# # Session & Auth Handling
# # --------------------------
# if "auth_token" not in st.session_state:
#     st.session_state.auth_token = None
# if "current_user" not in st.session_state:
#     st.session_state.current_user = None

# # Fetch username after login if token exists but username missing
# if not st.session_state.current_user and st.session_state.get("auth_token"):
#     try:
#         headers = {"Authorization": f"Bearer {st.session_state.auth_token}"}
#         resp = requests.get("http://127.0.0.1:8000/me", headers=headers)
#         if resp.status_code == 200:
#             st.session_state.current_user = resp.json().get("username")
#         else:
#             st.warning("⚠️ Failed to fetch user info. Please re-login.")
#     except Exception as e:
#         st.error(f"Exception fetching user info: {e}")

# # --------------------------
# # Dynamic User Export Directory
# # --------------------------
# def get_user_exports_dir():
#     """Return the correct exports directory for the logged-in user."""
#     base_dir = os.path.join(PROJECT_ROOT, "exports")
#     username = st.session_state.get("current_user")
#     if username:
#         user_dir = os.path.join(base_dir, username)
#         os.makedirs(user_dir, exist_ok=True)
#         return user_dir
#     else:
#         os.makedirs(base_dir, exist_ok=True)
#         return base_dir

# EXPORTS_DIR = get_user_exports_dir()



# # --------------------------
# # Session State Defaults
# # --------------------------
# if "step" not in st.session_state:
#     st.session_state.step = 1
# if "uploaded_files" not in st.session_state:
#     st.session_state.uploaded_files = []
# if "uploaded_paths" not in st.session_state:
#     st.session_state.uploaded_paths = []
# if "jd_text" not in st.session_state:
#     st.session_state.jd_text = ""
# if "weights" not in st.session_state:
#     st.session_state.weights = {
#         "skills": 0.4,
#         "experience": 0.3,
#         "education": 0.2,
#         "certifications": 0.1
#     }

# # Ensure additional session keys exist to avoid Streamlit AttributeError
# if "evaluation_done" not in st.session_state:
#     st.session_state.evaluation_done = False
# if "evaluation_response" not in st.session_state:
#     st.session_state.evaluation_response = None
# if "results" not in st.session_state:
#     st.session_state.results = None
# if "excel_file" not in st.session_state:
#     st.session_state.excel_file = None
# if "jd_data" not in st.session_state:
#     st.session_state.jd_data = None
# if "auth_token" not in st.session_state:
#     st.session_state.auth_token = None
# if "current_user" not in st.session_state:
#     st.session_state.current_user = None
# if "llm_model" not in st.session_state:
#     # default model (can be overridden by sidebar)
#     st.session_state.llm_model = "gemini-2.5-flash"
# if "llm_api_key" not in st.session_state:
#     st.session_state.llm_api_key = None
# if "save_mode" not in st.session_state:
#     st.session_state.save_mode = None
# if "show_auth" not in st.session_state:
#     # When True, the sidebar shows the login/register form. Clear after login.
#     st.session_state.show_auth = True


# def safe_rerun():
#     """Try to trigger a Streamlit rerun in a backwards/forwards-compatible way.

#     - Prefer st.experimental_rerun() where available.
#     - Otherwise change a query param (st.experimental_set_query_params) to force a rerun.
#     - If neither exists, toggle a session flag as a non-crashing fallback.
#     """
#     try:
#         fn = getattr(st, "experimental_rerun", None)
#         if callable(fn):
#             fn()
#             return

#         # Prefer the newer query params API. Assigning to st.query_params
#         # will update the URL params and force a rerun in modern Streamlit.
#         try:
#             qp = getattr(st, "query_params", None)
#             if qp is not None:
#                 import time

#                 st.query_params = {"_rerun": int(time.time())}
#                 return
#         except Exception:
#             # fall through to session toggle fallback
#             pass
#     except Exception as e:
#         # don't raise to UI; fall back to session toggle
#         print(f"safe_rerun helper failed: {e}")

#     # best-effort fallback: toggle a lightweight session-state flag
#     st.session_state["_needs_rerun"] = not st.session_state.get("_needs_rerun", False)


# def force_rerun():
#     """More aggressive rerun: try multiple methods to force Streamlit to refresh.

#     Tries, in order:
#     - st.experimental_rerun()
#     - st.experimental_set_query_params (if available)
#     - assignment to st.query_params
#     - toggling a session-state flag
#     """
#     try:
#         fn = getattr(st, "experimental_rerun", None)
#         if callable(fn):
#             fn()
#             return
#     except Exception:
#         pass

#     # try assigning st.query_params
#     try:
#         import time

#         qp = getattr(st, "query_params", None)
#         if qp is not None:
#             st.query_params = {"_rerun": int(time.time())}
#             return
#     except Exception:
#         pass

#     # fallback: toggle a session flag
#     st.session_state["_needs_rerun"] = not st.session_state.get("_needs_rerun", False)


# username = st.session_state.get("current_user", "Guest")
# top_col1, top_col2 = st.columns([8, 1])
# with top_col1:
#     st.subheader(f"👋 Welcome, {username}!")
#     st.write(f"📁 Current Export Folder: {EXPORTS_DIR}")
# with top_col2:
#     if st.button("🔄", help="Refresh app"):
#         force_rerun()


# # --------------------------
# # Sidebar: Account + LLM Settings
# # --------------------------
# with st.sidebar:
#     st.header("Account")
#     st.markdown("---")
#     st.subheader("LLM Settings (optional)")
#     model_choices = {
#         "Gemini 2.5 Flash (fast, recommended)": "gemini-2.5-flash",
#         "Gemini 2.5 Flash Lite (higher rpd)": "gemini-2.5-flash-lite",
#         "Gemini 1.5 Pro (balanced)": "gemini-1.5-pro",
#         "Gemini 1.0 (smaller, cheaper)": "gemini-1.0",
#         "Custom model name": "custom"
#     }
#     model_label = st.selectbox("Choose model", list(model_choices.keys()), index=0, key="_llm_model_select")
#     selected_model_value = model_choices[model_label]
#     if selected_model_value == "custom":
#         custom_model = st.text_input("Custom model name", key="_llm_model_custom")
#         model_value = custom_model or "gemini-2.5-flash"
#     else:
#         model_value = selected_model_value

#     api_key_input = st.text_input("Model API Key (optional)", type="password", key="_llm_api_key")




#     # persist choices in session state so other UI actions pick them up
#     if model_value:
#         st.session_state.llm_model = model_value
#     if api_key_input:
#         st.session_state.llm_api_key = api_key_input

#     # Put account controls inside a sidebar container so we can explicitly clear it
#     account_container = st.container()

#     supports_on_submit = False
#     # Callback to handle auth form submission. Using on_submit ensures
#     # session_state is updated immediately and allows us to call experimental_rerun
#     def _handle_auth_submit():
#         auth_tab = st.session_state.get("auth_tab", "Login")
#         username = st.session_state.get("_auth_username", "")
#         password = st.session_state.get("_auth_password", "")
#         full_name = st.session_state.get("_auth_fullname", "")
#         try:
#             if auth_tab == "Register":
#                 resp = requests.post(
#                     "http://127.0.0.1:8000/register",
#                     params={"username": username, "password": password, "full_name": full_name},
#                     timeout=10,
#                 )
#                 if resp.status_code == 200:
#                     st.success("Registration successful. Please login.")
#                 else:
#                     st.error(f"Registration failed: {resp.text}")
#             else:
#                 resp = requests.post(
#                     "http://127.0.0.1:8000/login",
#                     params={"username": username, "password": password},
#                     timeout=10,
#                 )
#                 if resp.status_code == 200:
#                     token = resp.json().get("access_token")
#                     st.session_state.auth_token = token
#                     st.session_state.current_user = username
#                     st.session_state.show_auth = False
#                 else:
#                     st.error(f"Login failed: {resp.text}")
#         except Exception as e:
#             st.error(f"Auth request failed: {e}")

#     with account_container:
#         # Account controls
#         if not st.session_state.get("show_auth") and st.session_state.get("current_user"):
#             # user already logged in -> show signed-in view
#             st.markdown(f"**Signed in as:** {st.session_state.current_user}")
#             if st.button("Logout", key="btn_logout"):
#                 # clear session state and remove the container contents immediately
#                 st.session_state.auth_token = None
#                 st.session_state.current_user = None
#                 st.session_state.show_auth = True
#                 st.success("Logged out")
#                 # best-effort immediate rerun
#                 force_rerun()
#         else:
#             # Use plain widgets (not st.form) to avoid Streamlit form lifecycle quirks
#             # Widgets write directly to st.session_state keys so the handler can read them.
#             st.selectbox("Action", ("Login", "Register"), key="auth_tab")
#             st.text_input("Username", key="_auth_username")
#             st.text_input("Password", type="password", key="_auth_password")
#             st.text_input("Full name (register only)", key="_auth_fullname")
#             if st.button("Submit", key="btn_auth_submit"):
#                 _handle_auth_submit()
#                 # Trigger a rerun so the sidebar updates immediately
#                 force_rerun()
#             # LLM settings are picked from session_state keys set by the sidebar


# # --------------------------
# # Step 1: Upload Resumes
# # --------------------------
# if st.session_state.step >= 1:
#     st.header("Step 1: Upload Resume Files")
#     uploaded_files = st.file_uploader(
#         "Upload resumes (PDF/DOCX)", type=["pdf", "docx"], accept_multiple_files=True
#     )

#     if uploaded_files:
#         st.session_state.uploaded_files = uploaded_files
#         if st.button("Upload Resume(s)", key="btn_upload_resumes"):
#             files_data = [("files", (file.name, file.getvalue())) for file in uploaded_files]
#             headers = {}
#             if st.session_state.auth_token:
#                 headers["Authorization"] = f"Bearer {st.session_state.auth_token}"
#             # include llm settings if present
#             if st.session_state.get("llm_model"):
#                 headers["X-Model"] = st.session_state.get("llm_model")
#             if st.session_state.get("llm_api_key"):
#                 headers["X-Api-Key"] = st.session_state.get("llm_api_key")
#             response = requests.post("http://127.0.0.1:8000/upload_resumes_only", files=files_data, headers=headers)
#             if response.status_code == 200:
#                 st.session_state.uploaded_paths = response.json()["uploaded_paths"]
#                 st.success("✅ Resume(s) uploaded successfully!")
#                 st.session_state.step = 2
#             else:
#                 st.error(f"Failed to upload resumes: {response.text}")





# # --------------------------
# # Step 2: Job Description & Weights
# # --------------------------
# if st.session_state.step >= 2:
#     st.header("Step 2: Enter Job Description & Adjust Weights")

#     example_jd = """Title: Software Engineer
# Experience: 2+ years
# Skills: Python, FastAPI, MongoDB, Streamlit
# Location: Remote
# Description: Build and maintain backend services for AI-powered systems.
# """

#     # ---------------- Instructions ----------------
#     st.markdown(
#         """
#         💡 **Tip:** Enter the Job Description in **key–value** format (each on a new line):  
#         ```
#         Title: Software Engineer
#         Skills: Python, FastAPI
#         Experience: 2 years
#         ```
#         Then click **Update JD** after editing.
#         """
#     )

#     # ---------------- JD Input ----------------
#     jd_text = st.text_area(
#         "📝 Paste Job Description here (optional — leave empty to skip):",
#         value=st.session_state.get("jd_text", example_jd),
#         height=220,
#         key="jd_input"
#     )

#     # ---- Update Button (replaces Ctrl+Enter) ----
#     if st.button("✅ Update JD"):
#         st.session_state.jd_text = jd_text
#         st.success("Job Description updated successfully!")

#     # ----------- Dynamic Field Extraction -----------
#     import re

#     def extract_fields_from_jd(jd_text):
#         """Extract keys from JD lines like 'Skills:' or 'Experience:'"""
#         fields = []
#         for line in jd_text.splitlines():
#             match = re.match(r"^\s*([A-Za-z_ ]+)\s*:", line)
#             if match:
#                 key = match.group(1).strip().lower().replace(" ", "_")
#                 fields.append(key)
#         return list(dict.fromkeys(fields))  # remove duplicates

#     # Use the stored JD (updated only when button pressed)
#     current_jd = st.session_state.get("jd_text", "")
#     fields = extract_fields_from_jd(current_jd) if current_jd.strip() else []

#     # remove unwanted fields like "title"
#     fields = [f for f in fields if f.lower() != "title"]

#     if not fields and current_jd.strip():
#         st.info("No specific fields detected — default sliders will be used.")
#         fields = ["skills", "experience", "education", "certifications"]

#     if current_jd.strip():
#         st.subheader("Weights (adjust using sliders)")
#         weights = {}
#         for field in fields:
#             default_val = st.session_state.weights.get(field, 0.2) * 100  # convert to %
#             weights[field] = st.slider(
#                 f"{field.capitalize()} Weight (%)",
#                 0, 100, int(default_val), 5,
#                 key=f"w_{field}"
#             )

#         total = sum(weights.values())
#         st.markdown(f"**Total Weight Sum:** `{total}%`")

#         if total > 100:
#             st.warning("⚠️ Total weight exceeds 100%. Please adjust sliders.")
#         if total < 100:
#             st.warning("⚠️ Total weight is below 100%. Please adjust sliders.")


#         weights = {k: round(v / 100, 2) for k, v in weights.items()}
#         st.session_state.weights = weights
#     else:
#         st.session_state.weights = {}
#         st.info("ℹ️ No JD entered — the system will only parse resumes (no scoring).")

#     # --------------- Evaluate Button ---------------
#     if st.button("Evaluate Resume(s)", key="btn_evaluate"):
#         if not st.session_state.uploaded_paths:
#             st.warning("Please upload at least one resume before proceeding.")
#         else:
#             headers = {}
#             if st.session_state.auth_token:
#                 headers["Authorization"] = f"Bearer {st.session_state.auth_token}"
#             if st.session_state.get("llm_model"):
#                 headers["X-Model"] = st.session_state.get("llm_model")
#             if st.session_state.get("llm_api_key"):
#                 headers["X-Api-Key"] = st.session_state.get("llm_api_key")

#             if current_jd.strip():
#                 # ✅ JD Mode — send JD and weights
#                 payload = {
#                     "jd_text": current_jd,
#                     "weights": st.session_state.weights
#                 }

#                 response = requests.post(
#                     "http://127.0.0.1:8000/upload_jd",
#                     json=payload,
#                     headers=headers
#                 )

#                 if response.status_code == 200:
#                     st.session_state.jd_data = response.json().get("jd_data")
#                     st.success("✅ JD uploaded successfully!")
#                 else:
#                     st.error(f"Failed to upload JD: {response.text}")
#                     st.stop()
#             else:
#                 st.session_state.jd_data = None

#             st.session_state.step = 3
# # --------------------------
# # Step 3: Evaluate Resumes
# # --------------------------
# if st.session_state.step >= 3:
#     st.header("Step 3: Evaluating Resumes...")

#     if not st.session_state.evaluation_done:
#         with st.spinner("⏳ Evaluating resumes... This may take a few minutes."):
#             headers = {}
#             if st.session_state.auth_token:
#                 headers["Authorization"] = f"Bearer {st.session_state.auth_token}"
#             if st.session_state.get("llm_model"):
#                 headers["X-Model"] = st.session_state.get("llm_model")
#             if st.session_state.get("llm_api_key"):
#                 headers["X-Api-Key"] = st.session_state.get("llm_api_key")

#             payload = {
#                 "uploaded_paths": st.session_state.uploaded_paths,
#                 "jd_data": st.session_state.jd_data  # may be None
#             }

#             try:
#                 response = requests.post(
#                     "http://127.0.0.1:8000/evaluate_resumes",
#                     json=payload,
#                     headers=headers
#                 )
#                 st.session_state.evaluation_response = response
#                 st.session_state.evaluation_done = True
#             except Exception as e:
#                 st.error(f"⚠️ Request failed: {e}")
#                 st.stop()

#     # Handle backend response
#     response = st.session_state.evaluation_response
#     if response.status_code == 200:
#         result_json = response.json()
#         st.session_state.results = result_json.get("data", [])
#         jd_mode = result_json.get("jd_mode", "disabled")

#         if jd_mode == "enabled":
#             st.success("✅ Evaluation completed with JD scoring!")
#         else:
#             st.success("✅ Resume parsing completed (no JD provided).")

#         st.session_state.step = 4
#     else:
#         st.error(f"❌ Evaluation failed: {response.text}")

import os
import base64
import requests
import streamlit as st
from openpyxl import load_workbook
import re
import pymongo

# --------------------------
# Project Root
# --------------------------
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# --------------------------
# Session & Auth Handling
# --------------------------
if "auth_token" not in st.session_state:
    st.session_state.auth_token = None
if "current_user" not in st.session_state:
    st.session_state.current_user = None

# Fetch username after login if token exists but username missing
if not st.session_state.current_user and st.session_state.get("auth_token"):
    try:
        headers = {"Authorization": f"Bearer {st.session_state.auth_token}"}
        resp = requests.get("http://127.0.0.1:8000/me", headers=headers)
        if resp.status_code == 200:
            st.session_state.current_user = resp.json().get("username")
        else:
            st.warning("⚠️ Failed to fetch user info. Please re-login.")
    except Exception as e:
        st.error(f"Exception fetching user info: {e}")

# --------------------------
# Dynamic User Export Directory
# --------------------------
def get_user_exports_dir():
    """Return the correct exports directory for the logged-in user."""
    base_dir = os.path.join(PROJECT_ROOT, "exports")
    username = st.session_state.get("current_user")
    if username:
        user_dir = os.path.join(base_dir, username)
        os.makedirs(user_dir, exist_ok=True)
        return user_dir
    else:
        os.makedirs(base_dir, exist_ok=True)
        return base_dir

EXPORTS_DIR = get_user_exports_dir()



# --------------------------
# Session State Defaults
# --------------------------
if "step" not in st.session_state:
    st.session_state.step = 1
if "uploaded_files" not in st.session_state:
    st.session_state.uploaded_files = []
if "uploaded_paths" not in st.session_state:
    st.session_state.uploaded_paths = []
if "jd_text" not in st.session_state:
    st.session_state.jd_text = ""
if "weights" not in st.session_state:
    st.session_state.weights = {
        "skills": 0.4,
        "experience": 0.3,
        "education": 0.2,
        "certifications": 0.1
    }

# Ensure additional session keys exist to avoid Streamlit AttributeError
if "evaluation_done" not in st.session_state:
    st.session_state.evaluation_done = False
if "evaluation_response" not in st.session_state:
    st.session_state.evaluation_response = None
if "results" not in st.session_state:
    st.session_state.results = None
if "excel_file" not in st.session_state:
    st.session_state.excel_file = None
if "jd_data" not in st.session_state:
    st.session_state.jd_data = None
if "auth_token" not in st.session_state:
    st.session_state.auth_token = None
if "current_user" not in st.session_state:
    st.session_state.current_user = None
if "llm_model" not in st.session_state:
    # default model (can be overridden by sidebar)
    st.session_state.llm_model = "gemini-2.5-flash"
if "llm_api_key" not in st.session_state:
    st.session_state.llm_api_key = None
if "save_mode" not in st.session_state:
    st.session_state.save_mode = None
if "show_auth" not in st.session_state:
    # When True, the sidebar shows the login/register form. Clear after login.
    st.session_state.show_auth = True


def safe_rerun():
    """Try to trigger a Streamlit rerun in a backwards/forwards-compatible way.

    - Prefer st.experimental_rerun() where available.
    - Otherwise change a query param (st.experimental_set_query_params) to force a rerun.
    - If neither exists, toggle a session flag as a non-crashing fallback.
    """
    try:
        fn = getattr(st, "experimental_rerun", None)
        if callable(fn):
            fn()
            return

        # Prefer the newer query params API. Assigning to st.query_params
        # will update the URL params and force a rerun in modern Streamlit.
        try:
            qp = getattr(st, "query_params", None)
            if qp is not None:
                import time

                st.query_params = {"_rerun": int(time.time())}
                return
        except Exception:
            # fall through to session toggle fallback
            pass
    except Exception as e:
        # don't raise to UI; fall back to session toggle
        print(f"safe_rerun helper failed: {e}")

    # best-effort fallback: toggle a lightweight session-state flag
    st.session_state["_needs_rerun"] = not st.session_state.get("_needs_rerun", False)


def force_rerun():
    """More aggressive rerun: try multiple methods to force Streamlit to refresh.

    Tries, in order:
    - st.experimental_rerun()
    - st.experimental_set_query_params (if available)
    - assignment to st.query_params
    - toggling a session-state flag
    """
    try:
        fn = getattr(st, "experimental_rerun", None)
        if callable(fn):
            fn()
            return
    except Exception:
        pass

    # try assigning st.query_params
    try:
        import time

        qp = getattr(st, "query_params", None)
        if qp is not None:
            st.query_params = {"_rerun": int(time.time())}
            return
    except Exception:
        pass

    # fallback: toggle a session flag
    st.session_state["_needs_rerun"] = not st.session_state.get("_needs_rerun", False)


username = st.session_state.get("current_user") or "Guest"
top_col1, top_col2 = st.columns([8, 1])
with top_col1:
    st.subheader(f"👋 Welcome, {username}!")
    st.write(f"📁 Current Export Folder: {EXPORTS_DIR}")
with top_col2:
    if st.button("🔄", help="Refresh app"):
        force_rerun()


# --------------------------
# Sidebar: Account + LLM Settings
# --------------------------
with st.sidebar:
    st.header("Account")
    st.markdown("---")
    st.subheader("LLM Settings (optional)")
    model_choices = {
        "Gemini 2.5 Flash (fast, recommended)": "gemini-2.5-flash",
        "Gemini 2.5 Flash Lite (higher rpd)": "gemini-2.5-flash-lite",
        "Gemini 1.5 Pro (balanced)": "gemini-1.5-pro",
        "Gemini 1.0 (smaller, cheaper)": "gemini-1.0",
        "Custom model name": "custom"
    }
    model_label = st.selectbox("Choose model", list(model_choices.keys()), index=0, key="_llm_model_select")
    selected_model_value = model_choices[model_label]
    if selected_model_value == "custom":
        custom_model = st.text_input("Custom model name", key="_llm_model_custom")
        model_value = custom_model or "gemini-2.5-flash"
    else:
        model_value = selected_model_value

    api_key_input = st.text_input("Model API Key (optional)", type="password", key="_llm_api_key")




    # persist choices in session state so other UI actions pick them up
    if model_value:
        st.session_state.llm_model = model_value
    if api_key_input:
        st.session_state.llm_api_key = api_key_input

    # Put account controls inside a sidebar container so we can explicitly clear it
    account_container = st.container()

    supports_on_submit = False
    # Callback to handle auth form submission. Using on_submit ensures
    # session_state is updated immediately and allows us to call experimental_rerun
    # def _handle_auth_submit():
    #     auth_tab = st.session_state.get("auth_tab", "Login")
    #     username = st.session_state.get("_auth_username", "")
    #     password = st.session_state.get("_auth_password", "")
    #     full_name = st.session_state.get("_auth_fullname", "")
    #     try:
    #         if auth_tab == "Register":
    #             resp = requests.post(
    #                 "http://127.0.0.1:8000/register",
    #                 params={"username": username, "password": password, "full_name": full_name},
    #                 timeout=10,
    #             )
    #             if resp.status_code == 200:
    #                 st.success("Registration successful. Please login.")
    #             else:
    #                 st.error(f"Registration failed: {resp.text}")
    #         else:
    #             resp = requests.post(
    #                 "http://127.0.0.1:8000/login",
    #                 params={"username": username, "password": password},
    #                 timeout=10,
    #             )
    #             if resp.status_code == 200:
    #                 token = resp.json().get("access_token")
    #                 st.session_state.auth_token = token
    #                 st.session_state.current_user = username
    #                 st.session_state.show_auth = False
    #             else:
    #                 st.error(f"Login failed: {resp.text}")
    #     except Exception as e:
    #         st.error(f"Auth request failed: {e}")
    def _handle_auth_submit():
        auth_tab = st.session_state.get("auth_tab", "Login")
        username = st.session_state.get("_auth_username", "")
        password = st.session_state.get("_auth_password", "")
        full_name = st.session_state.get("_auth_fullname", "")

        try:
            if auth_tab == "Register":
                resp = requests.post(
                    "http://127.0.0.1:8000/register",
                    params={
                        "username": username,
                        "password": password,
                        "full_name": full_name
                    },
                    timeout=10,
                )

                try:
                    data = resp.json()
                except:
                    st.error("Server returned an invalid response.")
                    st.stop()

                if resp.status_code == 200:
                    st.success("Registration successful. Please login.")
                    st.stop()

                else:
                    msg = data.get("message", "Registration failed. Please try again.")
                    st.error(msg)
                    st.stop()

            # ---------------- LOGIN --------------------
            else:
                resp = requests.post(
                    "http://127.0.0.1:8000/login",
                    params={"username": username, "password": password},
                    timeout=10,
                )

                try:
                    data = resp.json()
                except:
                    st.error("Server returned an invalid response.")
                    st.stop()

                if resp.status_code == 200:
                    token = data.get("access_token")
                    if not token:
                        st.error("Login error: Missing token.")
                        st.stop()

                    st.session_state.auth_token = token
                    st.session_state.current_user = username
                    st.session_state.show_auth = False
                    st.success("Login successful!")
                    st.stop()

                elif resp.status_code == 401:
                    msg = data.get("message", "Incorrect username or password.")
                    st.error(msg)
                    st.stop()

                else:
                    msg = data.get("message", "Login failed. Please try again.")
                    st.error(msg)
                    st.stop()

        except requests.exceptions.Timeout:
            st.error("Authentication service timed out. Try again later.")
            st.stop()

        except requests.exceptions.ConnectionError:
            st.error("Auth server not reachable. Check if backend is running.")
            st.stop()

        except Exception:
            st.error("Unexpected error during authentication.")
            st.stop()


    with account_container:
        # Account controls
        if not st.session_state.get("show_auth") and st.session_state.get("current_user"):
            # user already logged in -> show signed-in view
            st.markdown(f"**Signed in as:** {st.session_state.current_user}")
            if st.button("Logout", key="btn_logout"):
                # clear session state and remove the container contents immediately
                st.session_state.auth_token = None
                st.session_state.current_user = None
                st.session_state.show_auth = True
                st.success("Logged out")
                # best-effort immediate rerun
                force_rerun()
        else:
            # Use plain widgets (not st.form) to avoid Streamlit form lifecycle quirks
            # Widgets write directly to st.session_state keys so the handler can read them.
            st.selectbox("Action", ("Login", "Register"), key="auth_tab")
            st.text_input("Username", key="_auth_username")
            st.text_input("Password", type="password", key="_auth_password")
            st.text_input("Full name (register only)", key="_auth_fullname")
            if st.button("Submit", key="btn_auth_submit"):
                _handle_auth_submit()
                # Trigger a rerun so the sidebar updates immediately
                force_rerun()
            # LLM settings are picked from session_state keys set by the sidebar


# --------------------------
# Step 1: MongoDB Credentials
# --------------------------

if st.session_state.step >= 1:
    st.header("Step 1: Enter MongoDB Details to Fetch Resumes")

    st.markdown(
        """
        Provide MongoDB connection details to fetch resumes (already parsed JSONs).
        """
    )

    mongo_url = st.text_input(
        "MongoDB URL",
        value=st.session_state.get(
            "mongo_url",
            "mongodb+srv://resume_db_user:Swapnil%4013@cluster0.kg6kzel.mongodb.net/?appName=Cluster0",
        ),
    )

    col1, col2 = st.columns(2)

    with col1:
        db_name = st.text_input(
            "Database Name", value=st.session_state.get("db_name", "resume_db_new")
        )
    with col2:
        collection_name = st.text_input(
            "Collection Name", value=st.session_state.get("collection_name", "resumes_1")
        )

    # if st.button("Connect & Continue"):
    #     if not mongo_url or not db_name or not collection_name:
    #         st.error("Please fill in all MongoDB fields.")
    #     else:
    #         if not st.session_state.get("auth_token"):
    #             st.error("You must be logged in to connect.")
    #         else:
    #             with st.spinner("Connecting to MongoDB..."):
    #                 headers = {
    #                     "Authorization": f"Bearer {st.session_state.auth_token}"
    #                 }
    #                 payload = {
    #                     "mongo_url": mongo_url,
    #                     "db_name": db_name,
    #                     "collection_name": collection_name,
    #                 }
    #                 try:
    #                     resp = requests.post(
    #                         "http://127.0.0.1:8000/connect_mongo",
    #                         json=payload,
    #                         headers=headers,
    #                         timeout=25,
    #                     )
    #                     if resp.status_code == 200:
    #                         data = resp.json()
    #                         resume_count = data.get("resume_count", 0)
    #                         st.session_state.resume_count = resume_count
    #                         st.session_state.mongo_url = mongo_url
    #                         st.session_state.db_name = db_name
    #                         st.session_state.collection_name = collection_name

    #                         st.success(f"Connected! Total resumes found: {resume_count}")
    #                         st.session_state.step = 2
    #                     elif resp.status_code == 401:
    #                         st.error("Unauthorized. Please login again.")
    #                     else:
    #                         st.error(f"Connection failed: {resp.text}")
    #                 except requests.exceptions.RequestException as e:
    #                     st.error(f"API request failed: {e}")
    if st.button("Connect & Continue"):
        # --- BASIC VALIDATION ---
        if not mongo_url or not db_name or not collection_name:
            st.warning("All fields are required.")
            st.stop()

        if not st.session_state.get("auth_token"):
            st.warning("Please login before connecting to MongoDB.")
            st.stop()

        # --- SPINNER ---
        with st.spinner("Connecting to MongoDB..."):
            try:
                headers = {"Authorization": f"Bearer {st.session_state.auth_token}"}
                payload = {
                    "mongo_url": mongo_url,
                    "db_name": db_name,
                    "collection_name": collection_name,
                }

                resp = requests.post(
                    "http://127.0.0.1:8000/connect_mongo",
                    json=payload,
                    headers=headers,
                    timeout=25,
                )

                try:
                    data = resp.json()
                except ValueError:
                    st.error("Server returned an invalid response.")
                    # ❌ DO NOT STOP HERE INSIDE SPINNER
                    pass

            except requests.exceptions.Timeout:
                st.error("Server is taking too long to respond.")
                pass
            except requests.exceptions.ConnectionError:
                st.error("Cannot reach backend API.")
                pass
            except Exception:
                st.error("Unexpected error occurred.")
                pass

        # --- AFTER SPINNER (SAFE ZONE) ---
        # SUCCESS
        if resp.status_code == 200:
            st.session_state.resume_count = data.get("resume_count", 0)
            st.session_state.mongo_url = mongo_url
            st.session_state.db_name = db_name
            st.session_state.collection_name = collection_name
            st.session_state.step = 2

            st.success(f"Connected! Resumes found: {st.session_state.resume_count}")
            safe_rerun()

        # AUTH ERROR
        elif resp.status_code == 401:
            st.session_state.auth_token = None
            st.session_state.step = 0
            st.error(data.get("message", "Session expired. Please login again."))
            st.stop()

        # OTHER BACKEND ERRORS
        else:
            st.error(data.get("message", "Unable to connect to MongoDB."))
            st.stop()




# --------------------------
# Step 2: Job Description & Weights
# --------------------------
if st.session_state.step >= 2:
    st.header("Step 2: Enter Job Description & Adjust Weights")

    example_jd = """Title: Software Engineer
Experience: 2+ years
Skills: Python, FastAPI, MongoDB, Streamlit
Location: Remote
Description: Build and maintain backend services for AI-powered systems.
"""

    st.markdown(
        """
        💡 **Tip:** Enter the Job Description in **key–value** format (each on a new line):  
        ```
        Title: Software Engineer
        Skills: Python, FastAPI
        Experience: 2 years
        ```
        Then click **Update JD** after editing.
        """
    )

    jd_text = st.text_area(
        "📝 Paste Job Description here:",
        value=st.session_state.get("jd_text", example_jd),
        height=220,
        key="jd_input"
    )

    if st.button("Update JD"):
        if jd_text.strip() == "":
            st.error("❌ Job Description cannot be empty.")
        else:
            st.session_state.jd_text = jd_text
            st.success("Job Description updated successfully!")

    

    def extract_fields_from_jd(jd_text):
        """Extract keys from JD lines like 'Skills:' or 'Experience:'"""
        fields = []
        for line in jd_text.splitlines():
            match = re.match(r"^\s*([A-Za-z_ ]+)\s*:", line)
            if match:
                key = match.group(1).strip().lower().replace(" ", "_")
                fields.append(key)
        return list(dict.fromkeys(fields))  # remove duplicates

    current_jd = st.session_state.get("jd_text", "")
    fields = extract_fields_from_jd(current_jd) if current_jd.strip() else []

    # remove unwanted fields like "title"
    fields = [f for f in fields if f.lower() != "title"]

    if not fields and current_jd.strip():
        st.info("No specific fields detected — default sliders will be used.")
        fields = ["skills", "experience", "education", "certifications"]

    if current_jd.strip():
        st.subheader("Weights (adjust using sliders)")
        weights = {}
        for field in fields:
            default_val = st.session_state.weights.get(field, 0.2) * 100  # convert to %
            weights[field] = st.slider(
                f"{field.capitalize()} Weight (%)",
                0, 100, int(default_val), 5,
                key=f"w_{field}"
            )

        total = sum(weights.values())
        st.markdown(f"**Total Weight Sum:** `{total}%`")

        if total > 100:
            st.warning("⚠️ Total weight exceeds 100%. Please adjust sliders.")
        if total < 100:
            st.warning("⚠️ Total weight is below 100%. Please adjust sliders.")

        weights = {k: round(v / 100, 2) for k, v in weights.items()}
        st.session_state.weights = weights
    else:
        st.session_state.weights = {}
        st.info("ℹ️ Please enter a valid Job Description to proceed.")

    # Disable Evaluate button if JD empty
    evaluate_disabled = not bool(current_jd.strip())

    if st.button("Evaluate Resume(s)", key="btn_evaluate", disabled=evaluate_disabled):
        headers = {}
        if st.session_state.auth_token:
            headers["Authorization"] = f"Bearer {st.session_state.auth_token}"
        if st.session_state.get("llm_model"):
            headers["X-Model"] = st.session_state.get("llm_model")
        if st.session_state.get("llm_api_key"):
            headers["X-Api-Key"] = st.session_state.get("llm_api_key")

        jd_data = {
            "jd_text": current_jd,
            "weights": st.session_state.weights
        }

        resp = requests.post("http://127.0.0.1:8000/upload_jd", json=jd_data, headers=headers)
        if resp.status_code == 200:
            st.session_state.jd_data = resp.json().get("jd_data")
            st.success("✅ JD uploaded successfully!")
            st.session_state.step = 3
        else:
            st.error(f"Failed to upload JD: {resp.text}")
            st.stop()


# --------------------------
# Step 3: Evaluate Resumes (from DB)
# --------------------------
if st.session_state.step >= 3:
    st.header("Step 3: Evaluating Resumes from MongoDB...")

    if not st.session_state.evaluation_done:
        with st.spinner("⏳ Evaluating resumes... This may take a few minutes."):
            headers = {}
            if st.session_state.auth_token:
                headers["Authorization"] = f"Bearer {st.session_state.auth_token}"
            if st.session_state.get("llm_model"):
                headers["X-Model"] = st.session_state.get("llm_model")
            if st.session_state.get("llm_api_key"):
                headers["X-Api-Key"] = st.session_state.get("llm_api_key")

            payload = {
                "mongo_url": st.session_state.mongo_url,
                "db_name": st.session_state.db_name,
                "collection_name": st.session_state.collection_name,
                "jd_data": st.session_state.jd_data  # may be None
            }

            try:
                response = requests.post(
                    "http://127.0.0.1:8000/evaluate_resumes_db",
                    json=payload,
                    headers=headers
                )
                st.session_state.evaluation_response = response
                st.session_state.evaluation_done = True
            except Exception as e:
                st.error(f"⚠️ Request failed: {e}")
                st.stop()

    # Process response
    response = st.session_state.evaluation_response
    if response.status_code == 200:
        result_json = response.json()
        st.session_state.results = result_json.get("data", [])
        jd_mode = result_json.get("jd_mode", "disabled")

        if jd_mode == "enabled":
            st.success("✅ Evaluation completed with JD scoring!")
        else:
            st.success("✅ Resume parsing completed (no JD provided).")

        st.session_state.step = 4
    else:
        st.error(f"❌ Evaluation failed: {response.text}")



# --------------------------
# Step 4: Display Results
# --------------------------
if st.session_state.step >= 4 and st.session_state.results:
    st.header("Step 4: Evaluation Results")

    for resume in st.session_state.results:
        name = resume.get("name", "Unnamed Candidate")
        formatted_name = name.title()
        with st.expander(f"{formatted_name}"):
            st.markdown(f"**Email:** {resume.get('email','')}")
            st.markdown(f"**Phone:** {resume.get('phone','')}")
            st.markdown(f"**Location:** {resume.get('location','')}")

            # URLs
            urls = resume.get("urls", [])
            if urls:
                st.markdown("**URLs:**")
                for url in urls:
                    st.markdown(f"- {url}")

            # Skills
            skills = resume.get("skills", [])
            if skills:
                st.markdown("**Skills:**")
                skill_cards = " ".join([
                    f"<span style='display:inline-block;background:#e0f7fa;color:#006064;"
                    f"padding:5px 10px;margin:3px;border-radius:8px'>{s}</span>"
                    for s in skills
                ])
                st.markdown(skill_cards, unsafe_allow_html=True)

            # Education
            education = resume.get("education", [])
            if education:
                st.markdown("**Education:**")
                edu_cards = " ".join([
                    f"<span style='display:inline-block;background:#f1f8e9;color:#33691e;"
                    f"padding:5px 10px;margin:3px;border-radius:8px'>"
                    f"{edu.get('degree','')} - {edu.get('institution','')} ({edu.get('year','')})</span>"
                    for edu in education if isinstance(edu, dict)
                ])
                st.markdown(edu_cards, unsafe_allow_html=True)

            # Experience
            experience = resume.get("experience", [])
            if experience:
                st.markdown("**Experience:**")
                exp_cards = " ".join([
                    f"<span style='display:inline-block;background:#fff3e0;color:#e65100;"
                    f"padding:5px 10px;margin:3px;border-radius:8px'>"
                    f"{exp.get('role','')} at {exp.get('company','')} "
                    f"({exp.get('start_date','')} - {exp.get('end_date','')})</span>"
                    for exp in experience if isinstance(exp, dict)
                ])
                st.markdown(exp_cards, unsafe_allow_html=True)

            # Projects
            projects = resume.get("projects", [])
            if projects:
                st.markdown("**Projects:**")
                proj_cards = " ".join([
                    f"<span style='display:inline-block;background:#f3e5f5;color:#4a148c;"
                    f"padding:5px 10px;margin:3px;border-radius:8px'>"
                    f"{proj.get('title','')} [{', '.join(proj.get('technologies',[]))}]: "
                    f"{proj.get('description','')}</span>"
                    for proj in projects if isinstance(proj, dict)
                ])
                st.markdown(proj_cards, unsafe_allow_html=True)

            # Certifications
            certs = resume.get("certifications", [])
            if certs:
                st.markdown("**Certifications:**")
                cert_cards = " ".join([
                    f"<span style='display:inline-block;background:#ede7f6;color:#311b92;"
                    f"padding:5px 10px;margin:3px;border-radius:8px'>{c}</span>"
                    for c in certs
                ])
                st.markdown(cert_cards, unsafe_allow_html=True)

            # --------------------------
            # Summary Section
            # --------------------------
            st.markdown("---")
            st.markdown("### 📊 Summary")

            # Matched Skills
            matched = resume.get("matched_skills", [])
            if matched:
                st.markdown("**✅ Matched Skills (from JD):**")
                matched_cards = " ".join([
                    f"<span style='display:inline-block;background:#c8e6c9;color:#1b5e20;"
                    f"padding:6px 12px;margin:3px;border-radius:10px'>{s}</span>"
                    for s in matched
                ])
                st.markdown(matched_cards, unsafe_allow_html=True)

            # Missing Skills
            missing = resume.get("missing_skills", [])
            if missing:
                st.markdown("**⚠️ Missing Skills (from JD):**")
                missing_cards = " ".join([
                    f"<span style='display:inline-block;background:#ffcdd2;color:#b71c1c;"
                    f"padding:6px 12px;margin:3px;border-radius:10px'>{s}</span>"
                    for s in missing
                ])
                st.markdown(missing_cards, unsafe_allow_html=True)

            # Other Skills
            other = resume.get("other_skills", [])
            if other:
                st.markdown("**🟡 Other Skills (not in JD):**")
                other_cards = " ".join([
                    f"<span style='display:inline-block;background:#fff9c4;color:#f57f17;"
                    f"padding:6px 12px;margin:3px;border-radius:10px'>{s}</span>"
                    for s in other
                ])
                st.markdown(other_cards, unsafe_allow_html=True)

            # Total Experience
            exp_years = resume.get("total_experience_years", "N/A")
            st.markdown(
                f"<span style='display:inline-block;background:#fff8e1;color:#f57f17;"
                f"padding:8px 14px;margin:3px;border-radius:10px;font-weight:600;'>"
                f"🧮 Total Experience: {exp_years}</span>",
                unsafe_allow_html=True
            )

            # Overall Score
            score = resume.get("score", "N/A")
            st.markdown(
                f"<span style='display:inline-block;background:#e8f5e9;color:#1b5e20;"
                f"padding:8px 14px;margin:3px;border-radius:10px;font-weight:600;'>"
                f"🎯 Overall Score: {score}</span>",
                unsafe_allow_html=True
            )

            # Remarks
            remarks = resume.get("remarks", [])
            if remarks:
                st.markdown("**💬 Remarks / Feedback:**")
                remark_cards = " ".join([
                    f"<span style='display:inline-block;background:#fce4ec;color:#880e4f;"
                    f"padding:6px 12px;margin:3px;border-radius:10px'>{r}</span>"
                    for r in remarks
                ])
                st.markdown(remark_cards, unsafe_allow_html=True)

            # Scoring Breakdown
            breakdown = resume.get("scoring_breakdown", {})
            if breakdown:
                st.markdown("**📈 Scoring Breakdown:**")
                breakdown_cards = " ".join([
                    f"<span style='display:inline-block;background:#e3f2fd;color:#0d47a1;"
                    f"padding:6px 12px;margin:3px;border-radius:10px'>{k.capitalize()}: {v}</span>"
                    for k, v in breakdown.items()
                ])
                st.markdown(breakdown_cards, unsafe_allow_html=True)


# --------------------------
# Step 5: Export Options
# --------------------------
if st.session_state.get("step", 0) >= 4 and st.session_state.get("results"):
    st.header("Step 5: Export Options")

    # --- default selection
    if "export_option" not in st.session_state:
        st.session_state.export_option = "Create New Excel File"

    # --- option radio
    st.session_state.export_option = st.radio(
        "Choose Export Option:",
        (
            "Create New Excel File",
            "Append to Existing Sheet",
            "Create New Sheet in Existing File",
            "Export to MongoDB Database",
        ),
    )

    # --------------------------
    # Utility to call backend
    # --------------------------
    def export_to_backend(params):
        try:
            headers = {}
            if st.session_state.get("auth_token"):
                headers["Authorization"] = f"Bearer {st.session_state.auth_token}"
            if st.session_state.get("llm_model"):
                headers["X-Model"] = st.session_state.llm_model
            if st.session_state.get("llm_api_key"):
                headers["X-Api-Key"] = st.session_state.llm_api_key

            response = requests.post(
                "http://127.0.0.1:8000/export_resumes_excel",
                json=params,
                headers=headers,
                timeout=120
            )

            if response.status_code == 200:
                resp_json = response.json()
                if resp_json.get("status") == "success" and resp_json.get("excel_file"):
                    st.session_state.excel_file = resp_json
                    st.success("✅ Export successful!")
                else:
                    st.warning(f"❌ Export failed: {resp_json.get('message', 'Unknown error')}")
            else:
                st.warning(f"❌ Backend returned status {response.status_code}")

        except Exception as e:
            st.error(f"❌ Exception during export: {e}")

    # --------------------------
    # 1️⃣  Create New Excel File
    # --------------------------
    if st.session_state.export_option == "Create New Excel File":
        st.session_state.save_mode = "new_file"
        file_name = st.text_input("Enter Excel File Name", "resumes.xlsx")
        sheet_name = st.text_input("Enter Sheet Name", "Sheet1")

        if st.button("Export New Excel", key="btn_export_new"):
            file_path = os.path.join(EXPORTS_DIR, file_name)
            params = {
                "processed_resumes": st.session_state.results,
                "mode": "new_file",
                "file_path": file_path,
                "sheet_name": sheet_name,
            }
            export_to_backend(params)

    # --------------------------
    # 2️⃣  Append to Existing Sheet
    # --------------------------
    elif st.session_state.export_option == "Append to Existing Sheet":
        st.session_state.save_mode = "append_sheet"
        excel_files = [f for f in os.listdir(EXPORTS_DIR) if f.endswith(".xlsx")]

        if excel_files:
            selected_file = st.selectbox("Select Existing Excel File", excel_files)
            if selected_file:
                file_path = os.path.join(EXPORTS_DIR, selected_file)
                wb = load_workbook(file_path)
                selected_sheet = st.selectbox("Select Existing Sheet", wb.sheetnames)

                if st.button("Append to Sheet", key="btn_append_sheet"):
                    params = {
                        "processed_resumes": st.session_state.results,
                        "mode": "append_sheet",
                        "file_path": file_path,
                        "sheet_name": selected_sheet,
                    }
                    export_to_backend(params)
        else:
            st.warning("No existing Excel files found in your exports folder.")

    # --------------------------
    # 3️⃣  Create New Sheet in Existing File
    # --------------------------
    elif st.session_state.export_option == "Create New Sheet in Existing File":
        st.session_state.save_mode = "new_sheet"
        excel_files = [f for f in os.listdir(EXPORTS_DIR) if f.endswith(".xlsx")]

        if excel_files:
            selected_file = st.selectbox("Select Existing Excel File", excel_files)
            if selected_file:
                file_path = os.path.join(EXPORTS_DIR, selected_file)
                new_sheet = st.text_input("Enter New Sheet Name", "Sheet1")

                if st.button("Create New Sheet", key="btn_create_sheet"):
                    params = {
                        "processed_resumes": st.session_state.results,
                        "mode": "new_sheet",
                        "file_path": file_path,
                        "sheet_name": new_sheet,
                    }
                    export_to_backend(params)
        else:
            st.warning("No existing Excel files found in your exports folder.")

    # --------------------------
    # 4️⃣  Export to MongoDB
    # --------------------------
    # elif st.session_state.export_option == "Export to MongoDB Database":
    #     st.session_state.save_mode = "mongo"
    #     mongo_url = st.text_input("MongoDB Connection URL", "")
    #     db_name = st.text_input("Database Name", "resume_db")
    #     collection_name = st.text_input("Collection Name", "resumes")

    #     if st.button("Export to MongoDB", key="btn_export_mongo"):
    #         if not mongo_url:
    #             st.error("Please enter a valid MongoDB URL.")
    #         else:
    #             params = {
    #                 "processed_resumes": st.session_state.results,
    #                 "mongo_url": mongo_url,
    #                 "db_name": db_name,
    #                 "collection_name": collection_name,
    #             }

    #             headers = {}
    #             if st.session_state.get("auth_token"):
    #                 headers["Authorization"] = f"Bearer {st.session_state.auth_token}"

    #             with st.spinner("⏳ Uploading resumes to MongoDB..."):
    #                 try:
    #                     response = requests.post(
    #                         "http://127.0.0.1:8000/export_resumes_mongo",
    #                         json=params,
    #                         headers=headers,
    #                         timeout=120,
    #                     )
    #                     resp_json = response.json()
    #                     if resp_json.get("status") == "success":
    #                         st.success(f"✅ Exported {resp_json.get('inserted_count')} resumes to MongoDB.")
    #                     else:
    #                         st.warning(f"❌ Export failed: {resp_json.get('message', 'Unknown error')}")
    #                 except Exception as e:
    #                     st.error(f"❌ Exception during export: {e}")
    elif st.session_state.export_option == "Export to MongoDB Database":
        st.session_state.save_mode = "mongo"
        mongo_url = st.text_input("MongoDB Connection URL", "")
        db_name = st.text_input("Database Name", "resume_db")
        collection_name = st.text_input("Collection Name", "resumes")

        if st.button("Export to MongoDB", key="btn_export_mongo"):
            # Basic validation
            if not mongo_url or not db_name or not collection_name:
                st.warning("All fields are required.")
                st.stop()

            if not st.session_state.get("auth_token"):
                st.warning("Please login before exporting to MongoDB.")
                st.stop()

            resp = None
            data = {}

            headers = {}
            if st.session_state.get("auth_token"):
                headers["Authorization"] = f"Bearer {st.session_state.auth_token}"

            params = {
                "processed_resumes": st.session_state.results,
                "mongo_url": mongo_url,
                "db_name": db_name,
                "collection_name": collection_name,
            }

            with st.spinner("⏳ Uploading resumes to MongoDB..."):
                try:
                    resp = requests.post(
                        "http://127.0.0.1:8000/export_resumes_mongo",
                        json=params,
                        headers=headers,
                        timeout=120,
                    )
                    try:
                        data = resp.json()
                    except ValueError:
                        st.error("Server returned an invalid response.")
                        # Don't stop here, spinner block should finish
                except requests.exceptions.Timeout:
                    st.error("Server is taking too long to respond.")
                    st.stop()
                except requests.exceptions.ConnectionError:
                    st.error("Cannot reach backend API.")
                    st.stop()
                except Exception:
                    st.error("Unexpected error occurred.")
                    st.stop()

            # After spinner
            if resp is None:
                st.error("No response from server.")
                st.stop()

            if resp.status_code == 200:
                if data.get("status") == "success":
                    inserted_count = data.get("inserted_count", 0)
                    st.success(f"✅ Exported {inserted_count} resumes to MongoDB.")
                else:
                    st.warning(f"❌ Export failed: {data.get('message', 'Unknown error')}")
            elif resp.status_code == 401:
                st.session_state.auth_token = None
                st.session_state.step = 0
                st.error(data.get("message", "Session expired. Please login again."))
                st.stop()
            else:
                st.error(data.get("message", "Unable to export resumes to MongoDB."))
                st.stop()

    # --------------------------
    # 5️⃣  Download Button
    # --------------------------
    if st.session_state.get("excel_file") and st.session_state.excel_file.get("excel_file"):
        excel_b64 = st.session_state.excel_file["excel_file"]
        excel_bytes = base64.b64decode(excel_b64)
        saved_path = st.session_state.excel_file.get("saved_path", "resumes.xlsx")
        st.download_button(
            label="Download Excel",
            data=excel_bytes,
            file_name=os.path.basename(saved_path),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
